# update_excel.py
# Scarica i nuovi prezzi ETF da GitHub e aggiorna lo sheet "Raw Data"
# di Portafoglio_Vittorio_PRO.xlsx con i prezzi grezzi.
# I 3 fondi (LU1883307461, IT0001080446, LU1883328467) vengono mantenuti
# con forward fill — aggiornali manualmente quando scarichi i NAV.
#
# USO: python update_excel.py
# REQUISITI: pip install pandas openpyxl requests

import pandas as pd
import requests
from io import StringIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime

# ── Configurazione ────────────────────────────────────────────────────────────
GITHUB_USER   = "VitAluigi"
GITHUB_REPO   = "Download-Prices"
GITHUB_BRANCH = "main"
CSV_PATH      = "data/ETF_Prices.csv"
CSV_URL = f"https://raw.githubusercontent.com/{GITHUB_USER}/{GITHUB_REPO}/{GITHUB_BRANCH}/{CSV_PATH}"

EXCEL_PATH = r"C:\Users\valuigi\OneDrive - KPMG\Desktop\Worth\Portafoglio Vittorio PRO.xlsx"
SHEET_NAME = "Raw Data"

ETF_COLS  = ["IE00B4L5Y983", "LU3038520774", "IE00BMG6Z448",
             "IE00BZ0PKT83", "FR0013416716", "IE00BDBRDM35"]
FUND_COLS = ["LU1883307461", "IT0001080446", "LU1883328467"]
ALL_COLS  = ["Date"] + ETF_COLS[:1] + FUND_COLS[:1] + FUND_COLS[1:2] + ETF_COLS[1:2] + \
            ETF_COLS[2:3] + ETF_COLS[3:4] + ETF_COLS[4:5] + FUND_COLS[2:] + ETF_COLS[5:]

# Ordine colonne uguale all'originale
COL_ORDER = ["Date", "IE00B4L5Y983", "LU1883307461", "IT0001080446",
             "LU3038520774", "IE00BMG6Z448", "IE00BZ0PKT83",
             "FR0013416716", "LU1883328467", "IE00BDBRDM35"]

# ── 1. Scarica nuovi prezzi da GitHub ────────────────────────────────────────
print("1. Downloading ETF prices from GitHub...")
r = requests.get(CSV_URL, timeout=30)
r.raise_for_status()
new_prices = pd.read_csv(StringIO(r.text), index_col=0, parse_dates=True)
new_prices.index.name = "Date"
print(f"   {len(new_prices)} rows, last date: {new_prices.index.max().date()}")

# ── 2. Carica Raw Data esistente ─────────────────────────────────────────────
print(f"\n2. Loading Raw Data from Excel...")
existing = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, header=0)

# Raw Data ha intestazioni in riga 1 (ISIN) e dati dalla riga 2
# Leggi solo le colonne con ISIN come header
existing.columns = [str(c).strip() for c in existing.columns]

# Trova la colonna Date
date_col = existing.columns[0]
existing[date_col] = pd.to_datetime(existing[date_col], errors='coerce')
existing = existing.dropna(subset=[date_col])
existing = existing.set_index(date_col)
existing.index.name = "Date"
existing = existing.sort_index()

last_date = existing.index.max()
print(f"   Last date in Raw Data: {last_date.date()}")

# ── 3. Trova nuove date ───────────────────────────────────────────────────────
new_dates = new_prices[new_prices.index > last_date]

if new_dates.empty:
    print("\nNessuna nuova riga da aggiungere. Raw Data già aggiornato.")
    exit()

print(f"\n3. Nuove date: {len(new_dates)} righe ({new_dates.index.min().date()} → {new_dates.index.max().date()})")

# ── 4. Costruisci nuove righe ─────────────────────────────────────────────────
new_rows = pd.DataFrame(index=new_dates.index)

# ETF: prezzi da GitHub
for isin in ETF_COLS:
    if isin in new_dates.columns:
        new_rows[isin] = new_dates[isin]

# Fondi: forward fill dall'ultimo valore disponibile in Raw Data
for fund in FUND_COLS:
    if fund in existing.columns:
        last_val = existing[fund].dropna()
        new_rows[fund] = last_val.iloc[-1] if len(last_val) > 0 else None
    else:
        new_rows[fund] = None

# ── 5. Unisci ─────────────────────────────────────────────────────────────────
updated = pd.concat([existing, new_rows]).sort_index()
updated = updated[~updated.index.duplicated(keep="last")]

# Mantieni solo le colonne originali nell'ordine corretto
cols_available = [c for c in COL_ORDER[1:] if c in updated.columns]
updated = updated[cols_available]

print(f"\n4. Totale righe: {len(updated)} | Ultima data: {updated.index.max().date()}")
print(updated[["IE00B4L5Y983", "LU1883307461", "IE00BDBRDM35"]].tail(5).to_string())

# ── 6. Scrivi in Excel preservando le altre celle/formule ────────────────────
print(f"\n5. Saving to Excel...")
wb = load_workbook(EXCEL_PATH)
ws = wb[SHEET_NAME]

# Trova la riga di header (dove ci sono gli ISIN)
header_row = None
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    if any(str(v) in ETF_COLS for v in row if v):
        header_row = i
        break

if header_row is None:
    print("WARN: header ISIN non trovato, uso riga 1")
    header_row = 1

# Leggi l'ordine delle colonne dall'header
header_values = [cell.value for cell in ws[header_row]]
col_map = {str(v): j+1 for j, v in enumerate(header_values) if v}

# Scrivi solo le nuove righe a partire dalla prima riga vuota
# (last_date + 1 in poi)
data_start_row = header_row + 1

# Trova l'ultima riga occupata
last_data_row = data_start_row
for row in ws.iter_rows(min_row=data_start_row, values_only=True):
    if any(v is not None for v in row):
        last_data_row += 1
    else:
        break

print(f"   Header row: {header_row}, writing from row: {last_data_row}")

# Scrivi le nuove righe
date_col_idx = col_map.get("Date", 1)
for date, row_data in new_rows.iterrows():
    ws.cell(row=last_data_row, column=date_col_idx, value=date.date())
    for isin, val in row_data.items():
        if isin in col_map and pd.notna(val):
            ws.cell(row=last_data_row, column=col_map[isin], value=round(float(val), 6))
    last_data_row += 1

wb.save(EXCEL_PATH)
print(f"\nDone. Excel salvato: {EXCEL_PATH}")
print(f"Righe aggiunte: {len(new_rows)}")
